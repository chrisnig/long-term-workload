from collections import namedtuple
from typing import Iterable
import os
import openpyxl
import openpyxl.cell
import openpyxl.chart
import openpyxl.styles
import openpyxl.worksheet
import util.io


class SatisfactionEvaluator:
    ModeInfo = namedtuple('ModeInfo', ['name', 'marker'])

    mode_settings = {
        "unequal": ModeInfo('unequal', 'diamond'),
        "equal": ModeInfo('equal', 'x'),
        "unequal-unfair": ModeInfo('unequal-unfair', 'diamond'),
        "equal-unfair": ModeInfo('equal-unfair', 'x')
    }

    colors = {
        "ASV": "ff5454",
        "APS": "ffa7a7",
        "ALV": "3ab7ff",
        "APL": "a3ddff"
    }

    satisfaction_table_cols = [(1, "avg sigma", 1, "AVERAGE", None, colors["APS"]),
                               (2, "var sigma", 1, "_xlfn.VAR.P", colors["ASV"], None),
                               (3, "avg l", 2, "AVERAGE", None, None),
                               (4, "var l", 2, "_xlfn.VAR.P", None, None),
                               (5, "avg lambda", 3, "AVERAGE", None, colors["APL"]),
                               (6, "var lambda", 3, "_xlfn.VAR.P", colors["ALV"], None)]

    def __init__(self, data_dir: str, solution_dir: str, unfair: bool=False) -> None:
        self.data_dir = data_dir
        self.solution_dir = solution_dir
        self.solution_data = None
        self.physicians = None
        self.results = None

        self.solver_modes = ["unequal", "equal"]
        if unfair:
            self.solver_modes = [x + "-unfair" for x in self.solver_modes]

        self.analysis_sheet = None
        self.charts_sheet = None

        # tables on per-directory sheets
        self.sat_first_col = 1
        self.sat_first_row = 1
        self.req_first_col = 6
        self.req_first_row = 1
        self.delta_req_first_col = 9
        self.delta_req_first_row = 1
        self.single_phys_var_first_col = 21
        self.single_phys_var_first_row = 1

        # tables on analysis sheet
        self.start_sat_table_col = 1
        self.start_sat_table_row = 1
        self.start_dir_table_col = self.start_sat_table_col + 2 + (len(self.satisfaction_table_cols) *
                                                                   len(self.solver_modes))
        self.start_dir_table_row = 1
        # self.start_req_table_col = self.start_dir_table_col + len(self.solver_modes) + 2
        # self.start_req_table_row = 1
        # self.start_single_phys_table_col = self.start_req_table_col + (len(self.solver_modes) * 4) + 4
        self.start_single_phys_table_row = 1

    def _find_continuous_physicians(self) -> set:
        physicians = None
        for directory in filter(lambda x: os.path.isdir(os.path.join(self.data_dir, x)),
                                sorted(os.listdir(self.data_dir))):
            data = util.io.CmplCdatReader.read(os.path.join(self.data_dir, directory, "transformed.cdat"))
            if physicians is None:
                physicians = data.get_set("J")
            else:
                physicians &= data.get_set("J")

        self.physicians = physicians
        return physicians

    def _read(self) -> dict:
        physicians = sorted(self.physicians)
        self.solution_data = dict()
        result = dict()

        for directory in filter(lambda x: os.path.isdir(os.path.join(self.solution_dir, x)),
                                sorted(os.listdir(self.solution_dir))):
            dir_result = dict()
            result[directory] = dir_result
            parameters = util.io.CmplCdatReader.read(os.path.join(self.data_dir, directory, "transformed.cdat"))
            days = parameters.get_scalar("W_max") * 7

            dir_result["total"] = {
                "g_req_on": len(parameters.get_values("g_req_on")),
                "g_req_off": len(parameters.get_values("g_req_off"))
            }

            for phys in physicians:
                dir_result[phys] = {
                    "g_req_on": parameters.count_if_exists("g_req_on", (phys, None, None, None)),
                    "g_req_off": parameters.count_if_exists("g_req_off", (phys, None, None))
                }

            dir_result["modes"] = dict()
            for solver_mode in self.solver_modes:
                mode_result = dict()
                dir_result["modes"][solver_mode] = mode_result
                solution = util.io.CmplSolutionReader.read(os.path.join(
                    self.solution_dir, directory, solver_mode + ".sol"))

                for phys in physicians:
                    req_on = parameters.count_if_exists("g_req_on", (phys, None, None, None))
                    req_off = parameters.count_if_exists("g_req_off", (phys, None, None))
                    delta_req_on = solution.count_if_exists("delta_req_on", (phys, None, None, None))
                    delta_req_off = solution.count_if_exists("delta_req_off", (phys, None, None))
                    sat = (req_on - delta_req_on + req_off - delta_req_off) / days
                    load = solution.get("l", (phys,))
                    varlambda = solution.count_if_exists("x", (phys, None, None, None)) / days

                    mode_result[phys] = {
                        "satisfaction": sat,
                        "delta_req_on": delta_req_on,
                        "delta_req_off": delta_req_off,
                        "load": load,
                        "lambda": varlambda
                    }

        self.results = result
        return result

    def _write(self, filename: str):
        wb = openpyxl.Workbook()
        wb.remove_sheet(wb.get_active_sheet())

        for directory in self.results:
            sheet = wb.create_sheet(directory)

            self._write_directory_satisfaction_table(sheet, directory)
            # self._write_directory_request_table(sheet, directory)
            # self._write_directory_delta_request_table(sheet, directory)

        analysis_sheet = wb.create_sheet("analysis", 0)
        self.analysis_sheet = analysis_sheet
        analysis_sheet.freeze_panes = "A2"
        charts_sheet = wb.create_sheet("charts", 0)
        self.charts_sheet = charts_sheet

        self._write_satisfaction_table(analysis_sheet, self.start_sat_table_col, self.start_sat_table_row)
        self._write_satisfaction_directory_table(analysis_sheet, self.start_dir_table_col, self.start_dir_table_row)
        # self._write_request_table(analysis_sheet, self.start_req_table_col, self.start_req_table_row)
        # self._write_single_physician_table(analysis_sheet, self.start_single_phys_table_col,
        #                                    self.start_single_phys_table_row)
        #
        # self._write_satisfaction_average_per_phys_chart(charts_sheet, 1, 1)
        # self._write_satisfaction_variance_per_phys_chart(charts_sheet, 11, 1)
        # self._write_single_physician_chart(charts_sheet, self.single_phys_var_first_col, self.single_phys_var_first_row)
        self._write_avg_sigma_chart(charts_sheet, 1, 1)
        self._write_avg_lambda_chart(charts_sheet, 11, 1)
        self._write_var_sigma_chart(charts_sheet, 1, 19)
        self._write_var_lambda_chart(charts_sheet, 11, 19)

        wb.save(filename)

    @staticmethod
    def _setBoldFont(cell: openpyxl.cell.Cell):
        cell.font = openpyxl.styles.Font(bold=True)

    @staticmethod
    def _setTopBorder(cell: openpyxl.cell.Cell):
        cell.border = openpyxl.styles.Border(top=openpyxl.styles.Side(border_style="thin", color="000000"))

    @staticmethod
    def _setCellBackground(cell: openpyxl.cell.Cell, color: str):
        cell.fill = openpyxl.styles.PatternFill("solid", fgColor=color)

    def _write_satisfaction_table(self, sheet: openpyxl.worksheet.Worksheet, first_col: int, first_row: int) -> None:
        cell = sheet.cell(column=first_col, row=first_row)
        cell.value = "phys"
        self._setBoldFont(cell)

        self._write_physician_table(sheet, first_col, first_row + 1)

        cell = sheet.cell(row=first_row + 1 + len(self.physicians), column=first_col)
        cell.value = "average"
        self._setBoldFont(cell)
        self._setTopBorder(cell)

        cell = sheet.cell(row=first_row + 2 + len(self.physicians), column=first_col)
        cell.value = "difference"
        self._setBoldFont(cell)

        cell = sheet.cell(row=first_row + 3 + len(self.physicians), column=first_col)
        cell.value = "diff %"
        self._setBoldFont(cell)

        cell = sheet.cell(row=first_row + 4 + len(self.physicians), column=first_col)
        cell.value = "variance"
        self._setBoldFont(cell)

        cell = sheet.cell(row=first_row + 5 + len(self.physicians), column=first_col)
        cell.value = "difference"
        self._setBoldFont(cell)

        cell = sheet.cell(row=first_row + 6 + len(self.physicians), column=first_col)
        cell.value = "diff %"
        self._setBoldFont(cell)

        for mode_cnt, solver_mode in enumerate(self.solver_modes):
            cols = self.satisfaction_table_cols
            for colinfo in cols:
                column = first_col + colinfo[0] + (mode_cnt * len(cols))
                cell = sheet.cell(row=first_row, column=column)
                cell.value = colinfo[1] + " " + solver_mode
                self._setBoldFont(cell)

                for phys_cnt in range(len(self.physicians)):
                    row = first_row + phys_cnt + 1

                    cell = sheet.cell(row=row, column=column)
                    dirsheet_row = self.sat_first_row + phys_cnt + 1
                    dirsheet_col = self.sat_first_col + (3 * mode_cnt) + colinfo[2]
                    dirsheet_cell = sheet.cell(row=dirsheet_row, column=dirsheet_col)
                    cell.value = "={}({})".format(colinfo[3], ",".join(["'{}'!{}".format(sheet_name,
                                                                                         dirsheet_cell.coordinate)
                                                                        for sheet_name in self.results]))

                last_phys_row = first_row + len(self.physicians)
                cell = sheet.cell(row=last_phys_row + 1, column=column)
                cell.value = "=AVERAGE({}:{})".format(
                     sheet.cell(row=first_row + 1, column=column).coordinate,
                     sheet.cell(row=last_phys_row, column=column).coordinate)
                self._setBoldFont(cell)
                self._setTopBorder(cell)
                if colinfo[4]:
                    self._setCellBackground(cell, colinfo[4])

                if mode_cnt > 0:
                    cell = sheet.cell(row=last_phys_row + 2, column=column)
                    cell.value = "={}-{}".format(
                        sheet.cell(row=last_phys_row + 1, column=column).coordinate,
                        sheet.cell(row=last_phys_row + 1, column=first_col + colinfo[0]).coordinate
                    )
                    self._setBoldFont(cell)
                    if colinfo[4]:
                        self._setCellBackground(cell, colinfo[4])

                    cell = sheet.cell(row=last_phys_row + 3, column=column)
                    cell.value = "={}/{}*100".format(
                        sheet.cell(row=last_phys_row + 2, column=column).coordinate,
                        sheet.cell(row=last_phys_row + 1, column=first_col + colinfo[0]).coordinate
                    )
                    self._setBoldFont(cell)
                if colinfo[4]:
                    self._setCellBackground(cell, colinfo[4])

                cell = sheet.cell(row=last_phys_row + 4, column=column)
                cell.value = "=_xlfn.VAR.P({}:{})".format(
                     sheet.cell(row=first_row + 1, column=column).coordinate,
                     sheet.cell(row=last_phys_row, column=column).coordinate)
                self._setBoldFont(cell)
                if colinfo[5]:
                    self._setCellBackground(cell, colinfo[5])

                if mode_cnt > 0:
                    cell = sheet.cell(row=last_phys_row + 5, column=column)
                    cell.value = "={}-{}".format(
                        sheet.cell(row=last_phys_row + 4, column=column).coordinate,
                        sheet.cell(row=last_phys_row + 4, column=first_col + colinfo[0]).coordinate
                    )
                    self._setBoldFont(cell)
                    if colinfo[5]:
                        self._setCellBackground(cell, colinfo[5])

                    cell = sheet.cell(row=last_phys_row + 6, column=column)
                    cell.value = "={}/{}*100".format(
                        sheet.cell(row=last_phys_row + 5, column=column).coordinate,
                        sheet.cell(row=last_phys_row + 4, column=first_col + colinfo[0]).coordinate
                    )
                    self._setBoldFont(cell)
                    if colinfo[5]:
                        self._setCellBackground(cell, colinfo[5])

        row = first_row + len(self.physicians) + 8
        col = first_col
        for color in self.colors:
            cell = sheet.cell(row=row, column=col)
            cell.value = color
            self._setCellBackground(cell, self.colors[color])
            col += 1

    def _write_satisfaction_directory_table(self, sheet: openpyxl.worksheet.Worksheet, first_col: int, first_row: int) \
            -> None:
        cell = sheet.cell(column=first_col, row=first_row)
        cell.value = "plan"
        self._setBoldFont(cell)

        row = first_row + 1
        for directory in self.results:
            cell = sheet.cell(column=first_col, row=row)
            cell.value = directory
            self._setBoldFont(cell)
            row += 1

        cell = sheet.cell(column=first_col, row=row)
        cell.value = "average"
        self._setBoldFont(cell)
        self._setTopBorder(cell)

        cell = sheet.cell(column=first_col, row=row + 1)
        cell.value = "difference"
        self._setBoldFont(cell)

        cell = sheet.cell(column=first_col, row=row + 2)
        cell.value = "diff %"
        self._setBoldFont(cell)

        mode_cnt = 1
        for solver_mode in self.solver_modes:
            column = first_col + mode_cnt
            cell = sheet.cell(column=column, row=first_row)
            cell.value = solver_mode
            self._setBoldFont(cell)
            row = first_row + 1
            for directory in self.results:
                sheet.cell(column=column, row=row).value = "=_xlfn.VAR.P('{}'!{}:{})".format(
                    directory,
                    sheet.cell(row=self.sat_first_row + 1, column=self.sat_first_col + mode_cnt).coordinate,
                    sheet.cell(row=self.sat_first_row + len(self.physicians),
                               column=self.sat_first_col + mode_cnt).coordinate
                )
                row += 1
            cell = sheet.cell(column=column, row=row)
            cell.value = "=AVERAGE({}:{})".format(
                sheet.cell(row=first_row, column=column).coordinate,
                sheet.cell(row=row - 1, column=column).coordinate
            )
            self._setBoldFont(cell)
            self._setTopBorder(cell)

            cell = sheet.cell(row=row + 1, column=column)
            cell.value = "={}-{}".format(
                sheet.cell(row=row, column=column).coordinate,
                sheet.cell(row=row, column=first_col + 1).coordinate)

            cell = sheet.cell(row=row + 2, column=column)
            cell.value = "={}/{}*100".format(
                sheet.cell(row=row + 1, column=column).coordinate,
                sheet.cell(row=row, column=first_col + 1).coordinate
            )

            mode_cnt += 1

    def _write_physician_table(self, sheet: openpyxl.worksheet.Worksheet, first_col: int, first_row: int) -> None:
        row = first_row
        for phys in sorted(self.physicians):
            cell = sheet.cell(row=row, column=first_col)
            cell.value = phys
            self._setBoldFont(cell)

            row += 1

    def _write_directory_satisfaction_table(self, sheet: openpyxl.worksheet.Worksheet, directory: str) -> None:
        first_col = self.sat_first_col
        first_row = self.sat_first_row
        cell = sheet.cell(column=first_col, row=first_row)
        cell.value = "phys"
        self._setBoldFont(cell)

        self._write_physician_table(sheet, first_col, first_row + 1)

        for mode_cnt, solver_mode in enumerate(self.solver_modes):
            first_mode_col = first_col + 1 + (3 * mode_cnt)
            cell = sheet.cell(column=first_mode_col, row=first_row)
            cell.value = "sigma " + solver_mode
            self._setBoldFont(cell)

            cell = sheet.cell(column=first_mode_col + 1, row=first_row)
            cell.value = "l " + solver_mode
            self._setBoldFont(cell)

            cell = sheet.cell(column=first_mode_col + 2, row=first_row)
            cell.value = "lambda " + solver_mode
            self._setBoldFont(cell)

            row = first_row + 1
            for phys in sorted(self.physicians):
                sheet.cell(row=row, column=first_mode_col).value = \
                    self.results[directory]["modes"][solver_mode][phys]["satisfaction"]
                sheet.cell(row=row, column=first_mode_col + 1).value = \
                    self.results[directory]["modes"][solver_mode][phys]["load"]
                sheet.cell(row=row, column=first_mode_col + 2).value = \
                    self.results[directory]["modes"][solver_mode][phys]["lambda"]
                row += 1

    def _write_directory_request_table(self, sheet: openpyxl.worksheet.Worksheet, directory: str) -> None:
        first_col = self.req_first_col
        first_row = self.req_first_row

        cell = sheet.cell(column=first_col, row=first_row)
        cell.value = "g_req_on"
        self._setBoldFont(cell)

        cell = sheet.cell(column=first_col + 1, row=first_row)
        cell.value = "g_req_off"
        self._setBoldFont(cell)

        for phys_cnt, phys in enumerate(sorted(self.physicians)):
            cell = sheet.cell(column=first_col, row=first_row + 1 + phys_cnt)
            cell.value = self.results[directory][phys]["g_req_on"]
            cell = sheet.cell(column=first_col + 1, row=first_row + 1 + phys_cnt)
            cell.value = self.results[directory][phys]["g_req_off"]

    def _write_directory_delta_request_table(self, sheet: openpyxl.worksheet.Worksheet, directory: str) -> None:
        first_col = self.delta_req_first_col
        first_row = self.delta_req_first_row

        for mode_cnt, mode in enumerate(self.solver_modes):
            req_on_col = first_col + (mode_cnt * 2)
            req_off_col = req_on_col + 1

            cell = sheet.cell(column=req_on_col, row=first_row)
            cell.value = mode + " delta_req_on"
            self._setBoldFont(cell)

            cell = sheet.cell(column=req_off_col, row=first_row)
            cell.value = mode + " delta_req_off"
            self._setBoldFont(cell)

            for phys_cnt, phys in enumerate(sorted(self.physicians)):
                cell = sheet.cell(column=req_on_col, row=first_row + 1 + phys_cnt)
                cell.value = self.results[directory]["modes"][mode][phys]["delta_req_on"]
                cell = sheet.cell(column=req_off_col, row=first_row + 1 + phys_cnt)
                cell.value = self.results[directory]["modes"][mode][phys]["delta_req_off"]

    def _write_request_table(self, sheet: openpyxl.worksheet.Worksheet, first_col: int, first_row: int) -> None:
        cell = sheet.cell(column=first_col, row=first_row)
        cell.value = "requests"
        self._setBoldFont(cell)

        total_col = first_col + 1
        cell = sheet.cell(column=total_col, row=first_row)
        cell.value = "total (continuous phys.)"
        self._setBoldFont(cell)

        overall_col = first_col + 2
        cell = sheet.cell(column=overall_col, row=first_row)
        cell.value = "total (all phys.)"
        self._setBoldFont(cell)

        for directory_cnt, directory in enumerate(self.results):
            dir_row = first_row + 1 + directory_cnt
            cell = sheet.cell(row=dir_row, column=first_col)
            cell.value = directory
            self._setBoldFont(cell)

            first_req_cell = sheet.cell(row=self.req_first_row + 1, column=self.req_first_col)
            last_req_cell = sheet.cell(row=self.req_first_row + len(self.physicians), column=self.req_first_col + 1)

            total_delta_cell = sheet.cell(row=dir_row, column=total_col)
            total_delta_cell.value = "=SUM('{}'!{}:{})".format(directory, first_req_cell.coordinate,
                                                               last_req_cell.coordinate)

            overall_cell = sheet.cell(row=dir_row, column=overall_col)
            overall_cell.value = self.results[directory]["total"]["g_req_on"] \
                                 + self.results[directory]["total"]["g_req_off"]

            for mode_cnt, mode in enumerate(self.solver_modes):
                first_delta_cell = sheet.cell(row=self.delta_req_first_row + 1,
                                              column=self.delta_req_first_col + (2 * mode_cnt))
                last_delta_cell = sheet.cell(row=self.delta_req_first_row + len(self.physicians),
                                             column=self.delta_req_first_col + (2 * mode_cnt) + 1)

                delta_overall_col = first_col + 3 + mode_cnt
                delta_cell = sheet.cell(row=dir_row, column=delta_overall_col)
                delta_cell.value = "=SUM('{}'!{}:{})".format(directory,
                                                             first_delta_cell.coordinate,
                                                             last_delta_cell.coordinate)

                percent_col = delta_overall_col + len(self.solver_modes)
                percent_cell = sheet.cell(row=dir_row, column=percent_col)
                percent_cell.value = "={}/{}*100".format(delta_cell.coordinate, total_delta_cell.coordinate)

                overall_delta_col = percent_col + len(self.solver_modes)
                overall_delta_cell = sheet.cell(row=dir_row, column=overall_delta_col)
                overall_delta_cell.value = self.results[directory]["modes"][mode]["total"]["delta_req_on"] \
                                           + self.results[directory]["modes"][mode]["total"]["delta_req_off"]

                overall_percent_col = overall_delta_col + len(self.solver_modes)
                overall_percent_cell = sheet.cell(row=dir_row, column=overall_percent_col)
                overall_percent_cell.value = "={}/{}*100".format(overall_delta_cell.coordinate, overall_cell.coordinate)

        row = first_row + len(self.results) + 1
        cell = sheet.cell(column=first_col, row=row)
        cell.value = "sum"
        self._setBoldFont(cell)
        self._setTopBorder(cell)

        total_sum_cell = sheet.cell(column=total_col, row=row)
        total_sum_cell.value = "=SUM({}:{})".format(sheet.cell(row=first_row + 1, column=total_col).coordinate,
                                                    sheet.cell(row=first_row + len(self.results),
                                                               column=total_col).coordinate)
        self._setBoldFont(total_sum_cell)
        self._setTopBorder(total_sum_cell)

        overall_sum_cell = sheet.cell(column=overall_col, row=row)
        overall_sum_cell.value = "=SUM({}:{})".format(sheet.cell(row=first_row + 1, column=overall_col).coordinate,
                                                      sheet.cell(row=first_row + len(self.results),
                                                                 column=overall_col).coordinate)
        self._setBoldFont(overall_sum_cell)
        self._setTopBorder(overall_sum_cell)

        for mode_cnt, mode in enumerate(self.solver_modes):
            delta_col = first_col + mode_cnt + 3
            cell = sheet.cell(column=delta_col, row=first_row)
            cell.value = "delta (continuous phys.) " + mode
            self._setBoldFont(cell)

            vio_col = delta_col + len(self.solver_modes)
            cell = sheet.cell(column=vio_col, row=first_row)
            cell.value = "%vio (continuous phys.) " + mode
            self._setBoldFont(cell)

            delta_overall_col = vio_col + len(self.solver_modes)
            cell = sheet.cell(column=delta_overall_col, row=first_row)
            cell.value = "delta (all phys.) " + mode
            self._setBoldFont(cell)

            vio_overall_col = delta_overall_col + len(self.solver_modes)
            cell = sheet.cell(column=vio_overall_col, row=first_row)
            cell.value = "%vio (all phys.) " + mode
            self._setBoldFont(cell)

            total_delta_cell = sheet.cell(column=delta_col, row=row)
            total_delta_cell.value = "=SUM({}:{})".format(sheet.cell(row=first_row + 1,
                                                                     column=delta_col).coordinate,
                                                          sheet.cell(row=first_row + len(self.results),
                                                                     column=delta_col).coordinate)
            self._setBoldFont(total_delta_cell)
            self._setTopBorder(total_delta_cell)

            percentage_cell = sheet.cell(column=vio_col, row=row)
            percentage_cell.value = "={}/{}*100".format(total_delta_cell.coordinate, total_sum_cell.coordinate)
            self._setBoldFont(percentage_cell)
            self._setTopBorder(percentage_cell)

            total_delta_overall_cell = sheet.cell(column=delta_overall_col, row=row)
            total_delta_overall_cell.value = "=SUM({}:{})".format(sheet.cell(row=first_row + 1,
                                                                             column=delta_overall_col).coordinate,
                                                                  sheet.cell(row=first_row + len(self.results),
                                                                             column=delta_overall_col).coordinate)
            self._setBoldFont(total_delta_overall_cell)
            self._setTopBorder(total_delta_overall_cell)

            total_vio_overall_cell = sheet.cell(column=vio_overall_col, row=row)
            total_vio_overall_cell.value = "={}/{}*100".format(total_delta_overall_cell.coordinate,
                                                               overall_sum_cell.coordinate)
            self._setBoldFont(total_vio_overall_cell)
            self._setTopBorder(total_vio_overall_cell)

    def _write_avg_sigma_chart(self, sheet: openpyxl.worksheet.Worksheet, first_col: int, first_row: int) -> None:
        self._add_physician_chart(
            sheet,
            first_col,
            first_row,
            "average sigma",
            "average sigma",
            1
        )

    def _write_avg_lambda_chart(self, sheet: openpyxl.worksheet.Worksheet, first_col: int, first_row: int) -> None:
        self._add_physician_chart(
            sheet,
            first_col,
            first_row,
            "average lambda",
            "average lambda",
            5
        )

    def _write_var_sigma_chart(self, sheet: openpyxl.worksheet.Worksheet, first_col: int, first_row: int) -> None:
        self._add_physician_chart(
            sheet,
            first_col,
            first_row,
            "variance sigma",
            "variance sigma",
            2
        )

    def _write_var_lambda_chart(self, sheet: openpyxl.worksheet.Worksheet, first_col: int, first_row: int) -> None:
        self._add_physician_chart(
            sheet,
            first_col,
            first_row,
            "variance lambda",
            "variance lambda",
            6
        )

    def _add_physician_chart(self, sheet: openpyxl.worksheet.Worksheet, first_col: int, first_row: int, title: str,
                             y_axis: str, data_col_offset: int) -> None:
        anchor = sheet.cell(column=first_col, row=first_row)
        cell = sheet.cell(column=first_col, row=first_row)
        cell.value = title

        chart = openpyxl.chart.LineChart()
        chart.style = 1
        chart.height = 9
        chart.width = 15
        chart.legend.position = "b"
        chart.x_axis.title = "physician"
        chart.y_axis.title = y_axis

        xref = openpyxl.chart.Reference(self.analysis_sheet, min_col=self.start_sat_table_col,
                                        max_col=self.start_sat_table_col,
                                        min_row=self.start_sat_table_row + 1,
                                        max_row=self.start_sat_table_row + len(self.physicians))
        chart.set_categories(xref)
        for mode_cnt, mode in enumerate(self.solver_modes):
            data_column = self.start_sat_table_col + data_col_offset + (len(self.satisfaction_table_cols) * mode_cnt)
            ref = openpyxl.chart.Reference(self.analysis_sheet, min_col=data_column, max_col=data_column,
                                           min_row=self.start_sat_table_row + 1,
                                           max_row=self.start_sat_table_row + len(self.physicians))
            mode_info = self.mode_settings[mode]
            series = openpyxl.chart.Series(ref, title=mode_info.name)
            series.marker.symbol = mode_info.marker
            series.graphicalProperties.line.noFill = True
            chart.append(series)

        sheet.add_chart(chart, anchor=anchor.offset(column=1).coordinate)

    def _write_single_physician_table(self, sheet: openpyxl.worksheet.Worksheet, first_col: int, first_row: int) \
            -> None:
        physician_data_cell = self.charts_sheet.cell(column=self.single_phys_var_first_col,
                                                     row=self.single_phys_var_first_row).offset(column=1)

        for mode_cnt, mode in enumerate(self.solver_modes):
            cell = sheet.cell(column=first_col + 1 + mode_cnt, row=first_row)
            cell.value = mode
            self._setBoldFont(cell)

        for dir_cnt, directory in enumerate(sorted(self.results)):
            row = first_row + 1 + dir_cnt

            header_cell = sheet.cell(row=row, column=first_col)
            header_cell.value = directory
            self._setBoldFont(header_cell)

            for mode_cnt, mode in enumerate(self.solver_modes):
                col = first_col + 1 + mode_cnt
                mode_col_letter = chr(ord('B') + mode_cnt)

                value_cell = sheet.cell(row=row, column=col)
                value_cell.value = "=INDIRECT(\"'\" & {} & \"'!{}\" & ('{}'!{} + 1))".format(
                    header_cell.coordinate, mode_col_letter, self.charts_sheet.title, physician_data_cell.coordinate)

    def _write_single_physician_chart(self, sheet: openpyxl.worksheet.Worksheet, first_col: int, first_row: int) \
            -> None:
        cell = sheet.cell(column=first_col, row=first_row)
        cell.value = "satisfaction for phys"

        cell.offset(column=1).value = 1

        chart = self._create_default_chart()
        chart.x_axis_title = "physician"
        chart.y_axis_title = "satisfaction"

        xref = openpyxl.chart.Reference(self.analysis_sheet, min_col=self.start_single_phys_table_col,
                                        max_col=self.start_single_phys_table_col,
                                        min_row=self.start_single_phys_table_row + 1,
                                        max_row=self.start_single_phys_table_row + len(self.results))
        chart.set_categories(xref)
        for mode_cnt, mode in enumerate(self.solver_modes):
            data_column = self.start_single_phys_table_col + 1 + mode_cnt
            ref = openpyxl.chart.Reference(self.analysis_sheet, min_col=data_column, max_col=data_column,
                                           min_row=self.start_single_phys_table_row + 1,
                                           max_row=self.start_single_phys_table_row + len(self.results))
            series = openpyxl.chart.Series(ref, title=self.mode_settings[mode].name)
            chart.append(series)

        sheet.add_chart(chart, anchor=cell.offset(column=1, row=1).coordinate)

    @staticmethod
    def _create_default_chart():
        chart = openpyxl.chart.LineChart()
        chart.style = 1
        chart.height = 9
        chart.height = 15
        chart.legend.position = "b"
        return chart

    def evaluate_to_file(self, filename: str):
        self._find_continuous_physicians()
        self._read()
        self._write(filename)
