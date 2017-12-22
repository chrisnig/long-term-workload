import openpyxl
import openpyxl.chart
from util.data import DataSet
from typing import Union


class CmplCdatReader:
    """Provides a converter from CMPL .cdat data files to DataSet instances.
    """

    @staticmethod
    def read(filename: str) -> DataSet:
        """
        Reads a .cdat file as used by CMPL for data input and outputs it as a DataSet instance.
        :param filename: The filename of the file to read
        :return: DataSet
        """

        data = DataSet()

        # Current read state
        # 0: not reading anything
        # 1: reading elements of a matrix
        state = 0
        var = None

        with open(filename, "r") as file:
            for line in file:
                if state == 0:
                    if "<" in line:
                        # find the type of the new variable
                        if "set" in line:
                            # our variable is a one-line set
                            # sets could probably also be multi-line, but we don't support those
                            var = line.split()[0][1:]
                            setData = line.split()[3:-1]
                            if ".." in setData[0]:
                                # we are dealing with a number range
                                limits = list(map(lambda x: int(x), setData[0].split("..")))
                                value = set(range(limits[0], limits[1]+1))
                            else:
                                # we are dealing with a space-separated list of numbers
                                value = set(map(lambda x: int(x), setData))
                            data.create_set(var, value)
                        elif "indices" in line:
                            # our variable is a multi-line matrix
                            state = 1
                            var = line.split("[")[0][1:]
                            dimensions = tuple(line.split("[")[1].split("]")[0].split(", "))
                            default = CmplCdatReader.tryParse(line.split("=")[1].split("indices")[0])
                            data.create(var, default, dimensions)
                        else:
                            # our variable is a scalar
                            var = line.split()[0][1:]
                            value = CmplCdatReader.tryParse(line.split(">")[0].split("<")[1])
                            data.create_scalar(var, value)
                elif state == 1:
                    if line.startswith(">"):
                        state = 0
                        var = None
                        continue

                    if var is None:
                        raise RuntimeError("Tried to read a variable without proper declaration!")

                    if line == "\n":
                        continue

                    values = line.split()
                    index = []
                    for entry in values[:-1]:
                        try:
                            index.append(int(entry))
                        except ValueError:
                            index.append(entry)
                    index = tuple(index)
                    value = CmplCdatReader.tryParse(values[-1:][0])
                    data.set(var, index, value)
        return data

    @staticmethod
    def tryParse(value: str) -> Union[int, float]:
        try:
            return int(value)
        except ValueError:
            return float(value)


class CmplSolutionReader:
    """Provides a converter for CMPL-generated .sol solution files to DataSet instances.
    """

    def __init__(self):
        pass

    @staticmethod
    def read(filename: str, variables: Union[None, iter]=None) -> DataSet:
        """
        Reads a CMPL-generated solution file and converts it to a DataSet
        :param filename: The filename of the file to read
        :param variables: An iterable containing the string names of variables to be read. None to read all variables.
        :return: DataSet
        """

        solution = DataSet()

        with open(filename, "r") as file:
            sep_count = 0
            for line in file:
                if line.startswith("-" * 105):
                    sep_count += 1
                    continue

                if sep_count < 3:
                    continue

                if sep_count >= 4:
                    break

                varWithIndex = line.split()[0]
                var = varWithIndex.split("[")[0]
                if variables is not None and var not in variables:
                    continue
                try:
                    value = int(line.split()[2])
                except ValueError:
                    value = float(line.split()[2])
                index = varWithIndex.split("[")[1][:-1]
                indexTuple = tuple(map(lambda x: int(x), index.split(",")))
                if var not in solution:
                    solution.create(var)
                solution.set(var, indexTuple, value)

        return solution


class HistoryWriter:
    def __init__(self,):
        self.wb = openpyxl.Workbook()
        self.wb.remove_sheet(self.wb.active)

    def add_sheet(self, name: str, histories: iter, step_histories: iter) -> None:
        sheet = self.wb.create_sheet(name)

        last_col = len(histories)

        chart = openpyxl.chart.LineChart()
        chart.style = 12
        chart.height = 7.11
        chart.width = 10.41
        chart.legend.position = "b"

        for column, particle in enumerate(histories):
            for row, value in enumerate(particle):
                sheet.cell(column=column+1, row=row+1).value = value
            values = openpyxl.chart.Reference(sheet, min_col=column+1, max_col=column+1, min_row=1,
                                              max_row=len(particle))
            series = openpyxl.chart.Series(values)
            series.spPr.ln.w = 28575
            series.spPr.ln.cap = "rnd"
            chart.append(series)

        chart_col = last_col + 2
        sheet.add_chart(chart, anchor=sheet.cell(row=1, column=chart_col).coordinate)

        first_step_col = chart_col + 7  # chart has a width of 7 columns

        for column, particle in enumerate(step_histories):
            for row, value in enumerate(particle):
                sheet.cell(column=first_step_col + column, row=row + 1).value = str(value)

    def save(self, filename: str):
        self.wb.save(filename)


class CmplCdatWriter:
    @staticmethod
    def write(data: DataSet, filename: str) -> None:
        with open(filename, "w") as file:
            # write scalars first, then sets, then matrices
            variables = data.get_variables_by_type(DataSet.TYPE_SCALAR)
            for name, variable in variables:
                file.write("%{} < {} >\n".format(name, variable["values"]))

            variables = data.get_variables_by_type(DataSet.TYPE_SET)
            for name, variable in variables:
                set_content = None

                # check if set is a range
                if len(variable["values"]) > 1:
                    set_max = max(variable["values"])
                    set_min = min(variable["values"])
                    if all(x in variable["values"] for x in range(set_min, set_max+1)):
                        set_content = "{}..{}".format(set_min, set_max)

                if not set_content:
                    set_content = " ".join(str(x) for x in variable["values"])

                file.write("%{} set < {} >\n".format(name, set_content))

            variables = data.get_variables_by_type(DataSet.TYPE_MATRIX)
            for name, variable in variables:
                file.write("%{}[{}] = {} indices <\n".format(
                    name, ", ".join(variable["dimensions"]), variable["default"]))

                # write some default string if our variable does not have any values set
                if not variable["values"]:
                    file.write("\t".join(tuple("0" for _ in variable["dimensions"]) + (str(variable["default"]),))
                               + "\n")
                else:
                    for entry, entry_value in variable["values"].items():
                        file.write("\t" + "\t".join(str(x) for x in (entry + (entry_value,))) + "\n")
                
                file.write(">\n")
